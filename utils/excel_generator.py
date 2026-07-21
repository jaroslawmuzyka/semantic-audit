import pandas as pd
import io
import zipfile
from utils.openai_llm import GapAnalysisResult, ContentScores, AuditReport
from utils.themes import get_theme
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

# Excel ma twardy limit 32 767 znaków na komórkę.
MAX_CELL_LEN = 32000


def _clean_cell(value):
    if isinstance(value, str):
        value = ILLEGAL_CHARACTERS_RE.sub("", value)
        if len(value) > MAX_CELL_LEN:
            value = value[:MAX_CELL_LEN] + "\n[... skrócono do limitu komórki Excela ...]"
    return value


def _sanitize_df(df: pd.DataFrame) -> pd.DataFrame:
    return df.apply(lambda col: col.map(_clean_cell))


def _long_text_df(column_name: str, text: str) -> pd.DataFrame:
    """Dzieli długi tekst na wiele wierszy, żeby nie uciąć go na limicie komórki."""
    text = text or ""
    chunks = [text[i:i + MAX_CELL_LEN] for i in range(0, len(text), MAX_CELL_LEN)] or [""]
    return pd.DataFrame({column_name: chunks})


def apply_premium_formatting(writer, theme_key: str = "PG"):
    theme = get_theme(theme_key)
    header_font = Font(name=theme["excel_font"], bold=True, color='FFFFFF')
    header_fill = PatternFill(
        start_color=theme["excel_header_fill"],
        end_color=theme["excel_header_fill"],
        fill_type='solid',
    )
    align_top = Alignment(vertical='top', wrap_text=True)
    align_center = Alignment(horizontal='center', vertical='center')

    for sheet_name in writer.sheets:
        ws = writer.sheets[sheet_name]

        # Freeze top row
        ws.freeze_panes = "A2"

        # Remove gridlines for Podsumowanie
        if sheet_name in ["Summary & CQS", "Podsumowanie"]:
            ws.sheet_view.showGridLines = False

        # Style headers and set column widths
        for col_idx, column in enumerate(ws.columns, 1):
            col_letter = get_column_letter(col_idx)
            header_cell = column[0]
            header_cell.font = header_font
            header_cell.fill = header_fill
            header_cell.alignment = align_center

            # Apply wrapping to all body cells
            max_len = 0
            for cell in column:
                if cell.row > 1:
                    cell.alignment = align_top

                # Colorize CQS Score if it exists
                if header_cell.value == "CQS Score" or header_cell.value == "CQS" or header_cell.value == "Wynik CQS":
                    if cell.row > 1 and isinstance(cell.value, (int, float)):
                        if cell.value >= 80:
                            cell.fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
                        elif cell.value < 50:
                            cell.fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
                        else:
                            cell.fill = PatternFill(start_color='FFF8E1', end_color='FFF8E1', fill_type='solid')

                try:
                    val_str = str(cell.value)
                    # Use a cap for max length to avoid too wide columns
                    line_len = max([len(line) for line in val_str.split('\n')])
                    if line_len > max_len:
                        max_len = line_len
                except:
                    pass

            # Smart width adjustment
            if max_len > 80:
                ws.column_dimensions[col_letter].width = 80
            elif max_len > 15:
                ws.column_dimensions[col_letter].width = max_len + 2
            else:
                ws.column_dimensions[col_letter].width = 15


def generate_excel_report(gap_analysis: GapAnalysisResult, scores: ContentScores, report: AuditReport, source_content: str, consolidated_competitors: str, theme_key: str = "PG") -> bytes:
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Sheet 1: Executive Summary
        summary_data = {
            "Metric": ["CQS Score", "AI Citability Score", "Podsumowanie"],
            "Value": [report.cqs_score, report.ai_citability_score, report.executive_summary]
        }
        df_summary = pd.DataFrame(summary_data)
        _sanitize_df(df_summary).to_excel(writer, sheet_name="Summary & CQS", index=False)

        # Sheet 2: Recommendations
        recs_data = []
        for r in report.recommendations:
            recs_data.append({
                "Priority": r.priority,
                "Title": r.title,
                "Context": r.context,
                "BEFORE": r.before_quote,
                "AFTER": r.after_generated,
                "Impact on CQS": f"+{r.impact_cqs}"
            })
        df_recs = pd.DataFrame(recs_data)
        _sanitize_df(df_recs).to_excel(writer, sheet_name="Recommendations", index=False)

        # Sheet 3: Competitor EAV Matrix
        eav_data = []
        for e in gap_analysis.eav_matrix:
            row = {
                "Attribute": e.attribute,
                "URR Type": e.urr_type,
                "Coverage": e.coverage,
                "Priority": e.priority,
                "Status": e.status
            }
            # Add K1-K10 columns
            for i in range(10):
                col_name = f"K{i+1}"
                if i < len(e.competitors_presence):
                    row[col_name] = "+" if e.competitors_presence[i] else "-"
                else:
                    row[col_name] = "Brak danych"
            eav_data.append(row)

        df_eav = pd.DataFrame(eav_data)
        _sanitize_df(df_eav).to_excel(writer, sheet_name="Competitor EAV Matrix", index=False)

        # Sheet 4: Scores
        scores_data = []
        for s in scores.dimensions.as_list():
            scores_data.append({
                "Dimension": s.dimension_name,
                "Score (0-10)": s.score,
                "Top Problem": s.top_problem,
                "Problematic Quote": s.before_quote
            })
        df_scores = pd.DataFrame(scores_data)
        _sanitize_df(df_scores).to_excel(writer, sheet_name="Dimension Scores", index=False)

        # Sheet 5: EEAT Signals
        eeat_data = []
        for e in scores.eeat_signals.as_list():
            eeat_data.append({
                "Dimension": e.dimension,
                "Score": e.score,
                "Present Signals": e.present_signals,
                "Missing Signals": e.missing_signals
            })
        df_eeat = pd.DataFrame(eeat_data)
        _sanitize_df(df_eeat).to_excel(writer, sheet_name="EEAT Analysis", index=False)

        # Sheet 6: Action Plan
        action_plan_data = {
            "Docelowa Struktura Nagłówków": [entry.heading for entry in report.target_structure] if getattr(report, "target_structure", None) else [],
            "Suggested BLUF (First Sentence)": [entry.bluf for entry in report.target_structure] if getattr(report, "target_structure", None) else []
        }

        df_action = pd.DataFrame(action_plan_data)
        _sanitize_df(df_action).to_excel(writer, sheet_name="Action Plan", index=False)

        # Sheet 7: Missing TF-IDF
        tfidf_data = {"Missing Terms": scores.missing_tf_idf_terms}
        df_tfidf = pd.DataFrame(tfidf_data)
        _sanitize_df(df_tfidf).to_excel(writer, sheet_name="Missing TF-IDF", index=False)

        # Sheet 8: Raw Source Content (dzielone na wiersze — limit komórki Excela)
        df_source = _long_text_df("Source Article Content", source_content)
        _sanitize_df(df_source).to_excel(writer, sheet_name="Raw Source Content", index=False)

        # Sheet 9: Raw Competitors Content
        df_comps = _long_text_df("Consolidated Competitors Content", consolidated_competitors)
        _sanitize_df(df_comps).to_excel(writer, sheet_name="Raw Competitors Content", index=False)

        apply_premium_formatting(writer, theme_key)

    output.seek(0)
    return output.getvalue()


def generate_master_excel_report(all_results: list, theme_key: str = "PG") -> bytes:
    # Sort results by CQS score ascending
    sorted_results = sorted(all_results, key=lambda x: x.get("report").cqs_score if x.get("report") else 100)

    full_data = []
    action_data = []
    eav_data = []

    total_cqs = 0
    total_ai = 0
    excellent_count = 0
    needs_improvement_count = 0
    total_articles = 0
    total_cost = 0.0
    total_tokens_in = 0
    total_tokens_out = 0

    for item in sorted_results:
        url = item.get("url", "")
        keyword = item.get("keyword", "")
        r = item.get("report")
        s = item.get("scores")
        g = item.get("gap_analysis")

        t_in = item.get("tokens_in", 0)
        t_out = item.get("tokens_out", 0)
        cost = item.get("cost", 0.0)

        if not r or not s or not g:
            continue

        total_articles += 1
        total_cqs += r.cqs_score
        total_ai += r.ai_citability_score
        total_cost += cost
        total_tokens_in += t_in
        total_tokens_out += t_out

        if r.cqs_score >= 80:
            excellent_count += 1
        else:
            needs_improvement_count += 1

        # --- 1. FULL ROW ---
        row_full = {
            "URL": url,
            "Fraza": keyword,
            "CQS Score": r.cqs_score,
            "AI Citability": r.ai_citability_score,
            "Executive Summary": r.executive_summary,
            "Missing TF-IDF": ", ".join(s.missing_tf_idf_terms) if s.missing_tf_idf_terms else ""
        }

        # Dimensions
        for dim in s.dimensions.as_list():
            row_full[f"{dim.dimension_name} Score"] = dim.score
            row_full[f"{dim.dimension_name} Problem"] = dim.top_problem

        # EEAT
        for eeat in s.eeat_signals.as_list():
            row_full[f"EEAT {eeat.dimension} Score"] = eeat.score
            row_full[f"EEAT {eeat.dimension} Missing"] = eeat.missing_signals

        # Recommendations
        for i, rec in enumerate(r.recommendations):
            row_full[f"Rec {i+1} Priority"] = rec.priority
            row_full[f"Rec {i+1} Title"] = rec.title
            row_full[f"Rec {i+1} Impact"] = f"+{rec.impact_cqs}"
            row_full[f"Rec {i+1} BEFORE"] = rec.before_quote
            row_full[f"Rec {i+1} AFTER"] = rec.after_generated

        full_data.append(row_full)

        # --- 3. ACTIONABLE ROW ---
        crit = [f"[- {rec.title} -]\nObecna treść:\n{rec.before_quote}\nPrzykładowa (nowa) treść:\n{rec.after_generated}" for rec in r.recommendations if rec.priority.upper() == "KRYTYCZNE"]
        high = [f"[- {rec.title} -]\nObecna treść:\n{rec.before_quote}\nPrzykładowa (nowa) treść:\n{rec.after_generated}" for rec in r.recommendations if rec.priority.upper() == "WYSOKIE"]
        med  = [f"[- {rec.title} -]\nObecna treść:\n{rec.before_quote}\nPrzykładowa (nowa) treść:\n{rec.after_generated}" for rec in r.recommendations if rec.priority.upper() == "ŚREDNIE"]

        structure = []
        if getattr(r, "target_structure", None):
            for entry in r.target_structure:
                structure.append(f"{entry.heading}\nPrzykładowy (nowy) BLUF: {entry.bluf}")

        eeat_miss = []
        for e in s.eeat_signals.as_list():
            if e.missing_signals and e.missing_signals.strip() != "":
                eeat_miss.append(f"[{e.dimension}]: {e.missing_signals}")

        row_action = {
            "URL": url,
            "Fraza": keyword,
            "CQS Score": r.cqs_score,
            "AI Citability": r.ai_citability_score,
            "Podsumowanie": r.executive_summary,
            "KRYTYCZNE Rekomendacje": "\n\n".join(crit) if crit else "Brak",
            "WYSOKIE Rekomendacje": "\n\n".join(high) if high else "Brak",
            "ŚREDNIE Rekomendacje": "\n\n".join(med) if med else "Brak",
            "Docelowa Struktura Nagłówków": "\n\n".join(structure) if structure else "Brak",
            "Braki E-E-A-T": "\n".join(eeat_miss) if eeat_miss else "Brak",
            "Brakujące Słowa (TF-IDF)": row_full["Missing TF-IDF"]
        }

        action_data.append(row_action)

        # --- 4. ZBIORCZY EAV MATRIX ---
        for e in g.eav_matrix:
            row_eav = {
                "URL": url,
                "Fraza": keyword,
                "Attribute": e.attribute,
                "URR Type": e.urr_type,
                "Coverage": e.coverage,
                "Priority": e.priority,
                "Status": e.status
            }
            for i in range(10):
                col_name = f"K{i+1}"
                if i < len(e.competitors_presence):
                    row_eav[col_name] = "+" if e.competitors_presence[i] else "-"
                else:
                    row_eav[col_name] = "Brak danych"
            eav_data.append(row_eav)

    # --- 5. PODSUMOWANIE ---
    avg_cqs = round(total_cqs / total_articles, 2) if total_articles > 0 else 0
    avg_ai = round(total_ai / total_articles, 2) if total_articles > 0 else 0
    summary_text = f"Sprawdzono {total_articles} artykułów. {excellent_count} z nich ma ocenę bardzo dobrą (CQS >= 80), {needs_improvement_count} jest do poprawy (CQS < 80). Średnia ilość punktów CQS to {avg_cqs}."

    summary_data = [
        {"Metryka": "Podsumowanie", "Wartość": summary_text},
        {"Metryka": "Zbadane adresy URL", "Wartość": total_articles},
        {"Metryka": "Średni wynik CQS", "Wartość": avg_cqs},
        {"Metryka": "Średni wynik AI Citability", "Wartość": avg_ai},
        {"Metryka": "Artykuły bardzo dobre (>=80)", "Wartość": excellent_count},
        {"Metryka": "Artykuły do poprawy (<80)", "Wartość": needs_improvement_count}
    ]

    df_full = _sanitize_df(pd.DataFrame(full_data))
    df_action = _sanitize_df(pd.DataFrame(action_data))
    df_eav = _sanitize_df(pd.DataFrame(eav_data))
    df_summary = _sanitize_df(pd.DataFrame(summary_data))

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_summary.to_excel(writer, sheet_name="Podsumowanie", index=False)
        df_action.to_excel(writer, sheet_name="Rekomendacje (Actionable)", index=False)
        df_eav.to_excel(writer, sheet_name="Zbiorczy Matrix EAV", index=False)
        df_full.to_excel(writer, sheet_name="Pełny Raport", index=False)

        apply_premium_formatting(writer, theme_key)

    output.seek(0)
    return output.getvalue()


def create_zip_archive(files_dict: dict) -> bytes:
    output = io.BytesIO()
    with zipfile.ZipFile(output, 'w', zipfile.ZIP_DEFLATED) as zf:
        for filename, filebytes in files_dict.items():
            zf.writestr(filename, filebytes)
    output.seek(0)
    return output.getvalue()
