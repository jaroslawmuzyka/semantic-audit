"""Przeregenerowanie i patchowanie WYŁĄCZNIE sekcji E-E-A-T w już istniejących
raportach (XLSX/HTML, pojedynczych i masowych).

Nie przebudowuje plików od zera — otwiera dokładnie ten plik, który użytkownik
wgrał, i nadpisuje tylko komórki/fragmenty HTML dotyczące E-E-A-T. Cała
reszta pliku (formatowanie, inne arkusze/sekcje, branding) zostaje bez zmian.
"""
import io
import re
import html as html_lib

import streamlit as st
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from utils.openai_llm import (
    analyze_eeat_only, EAVEntry, GapAnalysisResult, ScoreDimension, ContentScores,
    EEATDetail, EEATBreakdown, Recommendation, TargetHeading, AuditReport,
)
from utils.naming import safe_filename
from utils.themes import THEMES, THEME_KEYS

EEAT_ORDER = ["Experience", "Expertise", "Authority", "Trust"]

_ALIGN_TOP_WRAP = Alignment(vertical="top", wrap_text=True)


# --------------------------------------------------------------------------
# Detekcja typu pliku
# --------------------------------------------------------------------------

def detect_file_kind(filename: str, file_bytes: bytes) -> str:
    """Zwraca 'xlsx_single', 'xlsx_master', 'html_single', 'html_master' albo 'unknown'."""
    name = filename.lower()
    if name.endswith(".xlsx"):
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
        except Exception:
            return "unknown"
        sheets = set(wb.sheetnames)
        if "Rekomendacje (Actionable)" in sheets or "Pełny Raport" in sheets:
            return "xlsx_master"
        if "EEAT Analysis" in sheets:
            return "xlsx_single"
        return "unknown"
    if name.endswith(".html") or name.endswith(".htm"):
        text = file_bytes.decode("utf-8", errors="ignore")
        if 'class="url-details"' in text:
            return "html_master"
        if "Braki E-E-A-T" in text:
            return "html_single"
        return "unknown"
    return "unknown"


# --------------------------------------------------------------------------
# Ekstrakcja URL-i / treści źródłowej z wgranych plików
# --------------------------------------------------------------------------

def _unique_keep_order(items):
    seen = set()
    out = []
    for i in items:
        if i and i not in seen:
            seen.add(i)
            out.append(i)
    return out


def extract_urls_from_master_xlsx(file_bytes: bytes) -> list:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    for sheet_name in ("Rekomendacje (Actionable)", "Pełny Raport"):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        headers = list(next(rows, []))
        if "URL" not in headers:
            continue
        url_idx = headers.index("URL")
        urls = [str(r[url_idx]).strip() for r in rows if r and r[url_idx] and str(r[url_idx]).strip()]
        return _unique_keep_order(urls)
    return []


def extract_urls_from_master_html(file_bytes: bytes) -> list:
    text = file_bytes.decode("utf-8", errors="ignore")
    raw = re.findall(r'<span class="s-url">(.*?)</span>', text, re.DOTALL)
    return _unique_keep_order(html_lib.unescape(r).strip() for r in raw)


def extract_url_from_single_html(file_bytes: bytes) -> str:
    text = file_bytes.decode("utf-8", errors="ignore")
    m = re.search(r'<strong>URL:</strong>\s*<a href="([^"]*)"', text)
    return html_lib.unescape(m.group(1)).strip() if m else ""


def extract_keywords_from_master_xlsx(file_bytes: bytes) -> dict:
    """Lekki odczyt pary URL -> Fraza z master XLSX (bez budowania pełnych
    obiektów jak reconstruct_results_from_master_xlsx) — używany do podpowiedzi
    frazy przy pełnej ponownej analizie."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    for sheet_name in ("Pełny Raport", "Rekomendacje (Actionable)"):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        headers = list(next(rows, []))
        if "URL" not in headers or "Fraza" not in headers:
            continue
        url_idx, kw_idx = headers.index("URL"), headers.index("Fraza")
        return {
            str(r[url_idx]).strip(): str(r[kw_idx] or "").strip()
            for r in rows if r and r[url_idx]
        }
    return {}


def guess_url_for_individual_file(filename: str, known_urls: list) -> str:
    """Odgaduje URL dla pliku indywidualnego XLSX (który nie ma wbudowanego URL-a)
    na podstawie konwencji nazewnictwa 'audit_{safe_filename(url)}.xlsx' używanej
    przy eksporcie (patrz safe_filename() w utils/naming.py). Zwraca dopasowany
    URL albo pusty string, jeśli żaden ze znanych adresów nie pasuje do nazwy.
    """
    base = filename.rsplit("/", 1)[-1]
    stem = re.sub(r"\.(xlsx|html?)$", "", base, flags=re.IGNORECASE)
    if stem.startswith("audit_"):
        stem = stem[len("audit_"):]
    for url in known_urls:
        if safe_filename(url) == stem:
            return url
    return ""


def extract_source_content_from_single_xlsx(file_bytes: bytes) -> str:
    """Skleja z powrotem 'Raw Source Content' (dzielone na wiersze przy eksporcie limitem 32k)."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    if "Raw Source Content" not in wb.sheetnames:
        return ""
    ws = wb["Raw Source Content"]
    chunks = []
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row and row[0]:
            chunks.append(str(row[0]))
    return "".join(chunks)


# --------------------------------------------------------------------------
# Regeneracja przez OpenAI
# --------------------------------------------------------------------------

def regenerate_eeat(source_content, language, user_context, system_prompt, model_name):
    """Zwraca (lista dokładnie 4 EEATSignal w kolejności EEAT_ORDER, usage)."""
    breakdown, usage = analyze_eeat_only(source_content, system_prompt, language, user_context, model_name)
    return breakdown.as_list(), usage


def eeat_missing_lines(eeat_list) -> list:
    return [f"[{e.dimension}]: {e.missing_signals}" for e in eeat_list if e.missing_signals and e.missing_signals.strip()]


# --------------------------------------------------------------------------
# Patchowanie XLSX
# --------------------------------------------------------------------------

def _autofit_column(ws, col_idx, max_width=80, min_width=15):
    col_letter = get_column_letter(col_idx)
    max_len = 0
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=col_idx).value
        if val is None:
            continue
        line_len = max(len(line) for line in str(val).split("\n"))
        max_len = max(max_len, line_len)
    if max_len > max_width:
        ws.column_dimensions[col_letter].width = max_width
    elif max_len > min_width:
        ws.column_dimensions[col_letter].width = max_len + 2
    else:
        ws.column_dimensions[col_letter].width = min_width


def patch_single_xlsx(file_bytes: bytes, eeat_list: list) -> bytes:
    """Nadpisuje arkusz 'EEAT Analysis' (dokładnie 4 wiersze), reszta pliku bez zmian."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    if "EEAT Analysis" not in wb.sheetnames:
        raise ValueError("Nie znaleziono arkusza 'EEAT Analysis' w tym pliku.")
    ws = wb["EEAT Analysis"]

    if ws.max_row > 5:
        ws.delete_rows(6, ws.max_row - 5)

    for i, e in enumerate(eeat_list, start=2):
        ws.cell(row=i, column=1, value=e.dimension).alignment = _ALIGN_TOP_WRAP
        ws.cell(row=i, column=2, value=e.score).alignment = _ALIGN_TOP_WRAP
        ws.cell(row=i, column=3, value=e.present_signals).alignment = _ALIGN_TOP_WRAP
        ws.cell(row=i, column=4, value=e.missing_signals).alignment = _ALIGN_TOP_WRAP

    for col in range(1, 5):
        _autofit_column(ws, col)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def patch_master_xlsx(file_bytes: bytes, url_to_eeat: dict) -> bytes:
    """Nadpisuje tylko wiersze wskazane w url_to_eeat (klucz = URL) w:
    - 'Rekomendacje (Actionable)' -> kolumna 'Braki E-E-A-T'
    - 'Pełny Raport' -> kanoniczne kolumny 'EEAT {Wymiar} Score/Missing' (dodane, jeśli brakuje)
    Pozostałe wiersze i arkusze zostają bez zmian.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))

    if "Rekomendacje (Actionable)" in wb.sheetnames:
        ws = wb["Rekomendacje (Actionable)"]
        headers = [c.value for c in ws[1]]
        if "URL" in headers and "Braki E-E-A-T" in headers:
            url_col = headers.index("URL") + 1
            eeat_col = headers.index("Braki E-E-A-T") + 1
            for row in range(2, ws.max_row + 1):
                url_val = ws.cell(row=row, column=url_col).value
                key = str(url_val).strip() if url_val else ""
                if key in url_to_eeat:
                    lines = eeat_missing_lines(url_to_eeat[key])
                    cell = ws.cell(row=row, column=eeat_col, value="\n".join(lines) if lines else "Brak")
                    cell.alignment = _ALIGN_TOP_WRAP
            _autofit_column(ws, eeat_col)

    if "Pełny Raport" in wb.sheetnames:
        ws = wb["Pełny Raport"]
        headers = [c.value for c in ws[1]]
        if "URL" in headers:
            url_col = headers.index("URL") + 1
            header_style_src = ws.cell(row=1, column=1)

            col_index = {}
            next_col = ws.max_column
            for dim in EEAT_ORDER:
                for suffix in ("Score", "Missing"):
                    name = f"EEAT {dim} {suffix}"
                    if name in headers:
                        col_index[name] = headers.index(name) + 1
                    else:
                        next_col += 1
                        header_cell = ws.cell(row=1, column=next_col, value=name)
                        header_cell.font = header_style_src.font
                        header_cell.fill = header_style_src.fill
                        header_cell.alignment = header_style_src.alignment
                        col_index[name] = next_col
                        headers.append(name)

            for row in range(2, ws.max_row + 1):
                url_val = ws.cell(row=row, column=url_col).value
                key = str(url_val).strip() if url_val else ""
                if key not in url_to_eeat:
                    continue
                by_dim = {e.dimension: e for e in url_to_eeat[key]}
                for dim in EEAT_ORDER:
                    e = by_dim.get(dim)
                    if not e:
                        continue
                    sc = ws.cell(row=row, column=col_index[f"EEAT {dim} Score"], value=e.score)
                    sc.alignment = _ALIGN_TOP_WRAP
                    mc = ws.cell(row=row, column=col_index[f"EEAT {dim} Missing"], value=e.missing_signals)
                    mc.alignment = _ALIGN_TOP_WRAP

            for name, idx in col_index.items():
                _autofit_column(ws, idx)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# --------------------------------------------------------------------------
# Patchowanie HTML
# --------------------------------------------------------------------------

def _eeat_li_html(eeat_list, empty_text) -> str:
    lines = [
        f"<li>[{html_lib.escape(e.dimension)}]: {html_lib.escape(e.missing_signals)}</li>"
        for e in eeat_list if e.missing_signals and e.missing_signals.strip()
    ]
    return "".join(lines) if lines else empty_text


def patch_single_html(file_bytes: bytes, eeat_list: list) -> bytes:
    text = file_bytes.decode("utf-8")
    new_inner = _eeat_li_html(eeat_list, "<li>Brak istotnych braków w sygnałach E-E-A-T.</li>")

    marker_re = re.compile(
        r"(<!--EEAT_BLOCK:START-->\s*<ul>\s*)(.*?)(\s*</ul>\s*<!--EEAT_BLOCK:END-->)", re.DOTALL
    )
    if marker_re.search(text):
        text = marker_re.sub(lambda m: m.group(1) + new_inner + m.group(3), text, count=1)
    else:
        fallback_re = re.compile(r"(<h4>Braki E-E-A-T</h4>\s*<ul>\s*)(.*?)(\s*</ul>)", re.DOTALL)
        if not fallback_re.search(text):
            raise ValueError("Nie znaleziono sekcji 'Braki E-E-A-T' w tym pliku HTML.")
        text = fallback_re.sub(lambda m: m.group(1) + new_inner + m.group(3), text, count=1)

    return text.encode("utf-8")


def patch_master_html(file_bytes: bytes, url_to_eeat: dict) -> bytes:
    text = file_bytes.decode("utf-8")
    block_re = re.compile(r'<details class="url-details">.*?</details>', re.DOTALL)

    def patch_block(m):
        block = m.group(0)
        url_m = re.search(r'<span class="s-url">(.*?)</span>', block, re.DOTALL)
        if not url_m:
            return block
        key = html_lib.unescape(url_m.group(1)).strip()
        if key not in url_to_eeat:
            return block

        new_inner = _eeat_li_html(url_to_eeat[key], "<li>Brak</li>")
        esc_url = re.escape(html_lib.escape(key))
        marker_re = re.compile(
            r"(<!--EEAT_BLOCK:START:" + esc_url + r'-->\s*<ul class="data-list">\s*)(.*?)'
            r"(\s*</ul>\s*<!--EEAT_BLOCK:END:" + esc_url + r"-->)",
            re.DOTALL,
        )
        if marker_re.search(block):
            return marker_re.sub(lambda mm: mm.group(1) + new_inner + mm.group(3), block, count=1)

        fallback_re = re.compile(r'(<h4>Braki E-E-A-T:</h4>\s*<ul class="data-list">\s*)(.*?)(\s*</ul>)', re.DOTALL)
        return fallback_re.sub(lambda mm: mm.group(1) + new_inner + mm.group(3), block, count=1)

    new_text = block_re.sub(patch_block, text)
    return new_text.encode("utf-8")


# --------------------------------------------------------------------------
# Pełna ponowna analiza URL-a — rekonstrukcja pozostałych wierszy z master XLSX
# --------------------------------------------------------------------------
#
# W przeciwieństwie do patchowania E-E-A-T powyżej, pełna reanaliza zastępuje
# CAŁY wiersz danego URL-a świeżym wynikiem audytu. Zamiast ręcznie łatać
# komórki w wielowierszowej, dynamicznej strukturze kolumn, odtwarzamy z
# master XLSX-a pełne obiekty (GapAnalysisResult/ContentScores/AuditReport)
# dla WSZYSTKICH pozostałych, nietkniętych URL-i, podmieniamy wpis dla
# przeregenerowanego adresu na świeży wynik i budujemy plik na nowo tymi
# samymi funkcjami co żywy audyt (generate_master_excel_report/html).
#
# Pola nigdzie nie renderowane w eksporcie master (problematic_fragments,
# srl_patient_instances, before_quote dla wymiarów ogólnych, top_3_gaps_p1/
# root_attributes/unique_opportunities) i tak nie trafiają do żadnego arkusza
# ani sekcji HTML master — ich odtworzenie jako puste jest w pełni bezstratne.

def guess_theme_from_path(path: str, file_bytes: bytes = None) -> str:
    """Zgaduje branding (PG/WPP) pliku — najpierw po folderze w ścieżce (tak
    zapisuje je ta aplikacja), a w razie potrzeby po kolorze wypełnienia
    nagłówka arkusza (excel_header_fill z utils/themes.py)."""
    parts = path.replace("\\", "/").split("/")
    for tk in THEME_KEYS:
        if tk in parts:
            return tk

    if file_bytes:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
            header_cell = wb[wb.sheetnames[0]].cell(row=1, column=1)
            fill_rgb = (header_cell.fill.start_color.rgb or "").upper()
            for tk in THEME_KEYS:
                if THEMES[tk]["excel_header_fill"].upper() in fill_rgb:
                    return tk
        except Exception:
            pass

    return THEME_KEYS[0]


def _parse_target_structure(text) -> list:
    text = str(text or "").strip()
    if not text or text == "Brak":
        return []
    marker = "\nPrzykładowy (nowy) BLUF: "
    entries = []
    for chunk in text.split("\n\n"):
        if marker in chunk:
            heading, bluf = chunk.split(marker, 1)
        else:
            heading, bluf = chunk, ""
        entries.append(TargetHeading(heading=heading.strip(), bluf=bluf.strip()))
    return entries


@st.cache_data(show_spinner=False)
def reconstruct_results_from_master_xlsx(file_bytes: bytes) -> dict:
    """Odtwarza pełne dane (kształt jak wpisy mass_results) dla KAŻDEGO URL-a
    w danym master XLSX, na podstawie arkuszy 'Pełny Raport', 'Rekomendacje
    (Actionable)' i 'Zbiorczy Matrix EAV'. Zwraca dict {url: result_dict}.

    Cache'owane po treści pliku — to dość ciężki parsing, a strona Streamlit
    przelicza się od nowa przy każdej interakcji użytkownika.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    if "Pełny Raport" not in wb.sheetnames:
        return {}

    eav_by_url = {}
    if "Zbiorczy Matrix EAV" in wb.sheetnames:
        ws_eav = wb["Zbiorczy Matrix EAV"]
        rows_eav = ws_eav.iter_rows(values_only=True)
        headers_eav = list(next(rows_eav, []))
        idx_eav = {h: i for i, h in enumerate(headers_eav) if h is not None}
        if "URL" in idx_eav:
            for r in rows_eav:
                if not r:
                    continue
                u = str(r[idx_eav["URL"]] or "").strip()
                if not u:
                    continue
                presence = [
                    (idx_eav.get(f"K{i}") is not None and str(r[idx_eav[f"K{i}"]] or "").strip() == "+")
                    for i in range(1, 11)
                ]
                eav_by_url.setdefault(u, []).append(EAVEntry(
                    attribute=str(r[idx_eav.get("Attribute", -1)] or ""),
                    urr_type=str(r[idx_eav.get("URR Type", -1)] or ""),
                    coverage=str(r[idx_eav.get("Coverage", -1)] or ""),
                    competitors_presence=presence,
                    priority=str(r[idx_eav.get("Priority", -1)] or ""),
                    status=str(r[idx_eav.get("Status", -1)] or ""),
                ))

    structure_by_url = {}
    if "Rekomendacje (Actionable)" in wb.sheetnames:
        ws_action = wb["Rekomendacje (Actionable)"]
        rows_action = ws_action.iter_rows(values_only=True)
        headers_action = list(next(rows_action, []))
        idx_action = {h: i for i, h in enumerate(headers_action) if h is not None}
        if "URL" in idx_action and "Docelowa Struktura Nagłówków" in idx_action:
            for r in rows_action:
                if r:
                    structure_by_url[str(r[idx_action["URL"]] or "").strip()] = r[idx_action["Docelowa Struktura Nagłówków"]]

    ws_full = wb["Pełny Raport"]
    rows_full = ws_full.iter_rows(values_only=True)
    headers_full = list(next(rows_full, []))
    idx_full = {h: i for i, h in enumerate(headers_full) if h is not None}
    if "URL" not in idx_full:
        return {}

    dimension_names = sorted({h[:-len(" Problem")] for h in headers_full if h and h.endswith(" Problem")})
    rec_indices = sorted({
        int(m.group(1)) for h in headers_full if h
        for m in [re.match(r"^Rec (\d+) Priority$", h)] if m
    })

    results = {}
    for r in rows_full:
        if not r:
            continue
        url = str(r[idx_full["URL"]] or "").strip()
        if not url:
            continue

        def gv(name, default=None):
            i = idx_full.get(name)
            return r[i] if i is not None else default

        dimensions = [
            ScoreDimension(
                dimension_name=dim,
                score=int(gv(f"{dim} Score") or 0),
                top_problem=str(gv(f"{dim} Problem") or ""),
                before_quote="",
            )
            for dim in dimension_names
        ]

        eeat_kwargs = {}
        for dim, field in zip(EEAT_ORDER, ["experience", "expertise", "authority", "trust"]):
            eeat_kwargs[field] = EEATDetail(
                score=int(gv(f"EEAT {dim} Score") or 0),
                present_signals="",
                missing_signals=str(gv(f"EEAT {dim} Missing") or ""),
            )

        missing_tfidf = [t.strip() for t in str(gv("Missing TF-IDF") or "").split(",") if t.strip()]

        scores = ContentScores(
            dimensions=dimensions,
            problematic_fragments=[],
            srl_patient_instances=[],
            eeat_signals=EEATBreakdown(**eeat_kwargs),
            missing_tf_idf_terms=missing_tfidf,
        )

        recommendations = []
        for i in rec_indices:
            priority = gv(f"Rec {i} Priority")
            if not priority:
                continue
            impact_raw = str(gv(f"Rec {i} Impact") or "0").replace("+", "").strip()
            try:
                impact = int(impact_raw)
            except ValueError:
                impact = 0
            recommendations.append(Recommendation(
                priority=str(priority), title=str(gv(f"Rec {i} Title") or ""), context="",
                before_quote=str(gv(f"Rec {i} BEFORE") or ""), after_generated=str(gv(f"Rec {i} AFTER") or ""),
                impact_cqs=impact,
            ))

        report = AuditReport(
            cqs_score=int(gv("CQS Score") or 0),
            ai_citability_score=int(gv("AI Citability") or 0),
            executive_summary=str(gv("Executive Summary") or ""),
            recommendations=recommendations,
            target_structure=_parse_target_structure(structure_by_url.get(url, "")),
            eeat_ready_blocks="",
        )

        gap_analysis = GapAnalysisResult(
            eav_matrix=eav_by_url.get(url, []),
            top_3_gaps_p1=[], root_attributes=[], unique_opportunities=[],
        )

        results[url] = {
            "url": url, "keyword": str(gv("Fraza") or ""),
            "report": report, "scores": scores, "gap_analysis": gap_analysis,
            "tokens_in": 0, "tokens_out": 0, "cost": 0.0,
        }

    return results


def merge_fresh_results_into_master(reconstructed: dict, fresh_results: dict) -> list:
    """Łączy zrekonstruowane dane nietkniętych wierszy z nowymi, w pełni
    świeżymi wynikami (kształt jak zwraca utils.audit_pipeline.fetch_and_audit)
    dla przeregenerowanych URL-i. Zwraca listę gotową dla
    generate_master_excel_report / generate_master_html_report."""
    merged = dict(reconstructed)
    for url, fresh in fresh_results.items():
        merged[url] = {
            "url": url, "keyword": fresh["keyword"], "report": fresh["report"],
            "scores": fresh["scores"], "gap_analysis": fresh["gap_analysis"],
            "tokens_in": fresh["tokens_in"], "tokens_out": fresh["tokens_out"], "cost": fresh["cost"],
        }
    return list(merged.values())
